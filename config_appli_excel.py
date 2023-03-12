import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import pandas as pd
from datetime import datetime, timedelta


def make_excel_file(df):
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()

    # Sélectionner la feuille active
    ws = wb.active

    # Ajouter un titre à la feuille
    ws.title = "Recherche emploi"

    # Insérer les données dans la feuille à partir d'un DataFrame pandas
    #df = pd.read_csv('Data/Data_jobs.csv',delimiter='|',header=0)

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # Définir la taille des cellules
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length

    # Ajouter une couleur de fond aux cellules et aligner les cellules

    for row in ws.iter_rows(min_row=2, min_col=1):
        print(row[3].value)
        date_object = datetime.strptime(row[3].value, '%Y-%m-%d %H:%M:%S')
        
        if  datetime.now() - date_object < timedelta(days=7):
            for cell in row:
                # Couleurs
                cell.fill = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                # Alignement
                cell.alignment = Alignment(vertical='center')

    # Enregistrer le fichier Excel
    name_file = 'Exemple.xlsx'
    wb.save(name_file)


    return name_file

import subprocess
import os

# Chemin absolu du fichier Excel
excel_file_path = os.path.abspath('Exemple.xlsx')

# Ouvrir le fichier Excel avec l'application par défaut
subprocess.Popen([excel_file_path], shell=True)



